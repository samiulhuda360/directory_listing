from django.shortcuts import render, redirect
from .tasks import create_company_profile_post, test_post_to_wordpress, delete_from_wordpress, perform_test_task, delete_post_by_url, find_post_id_by_url, update_company_profile_post, post_summary_to_wordpress, get_root_domain
from django.shortcuts import render
import openpyxl
from .models import APIConfig, GeneratedURL, TestResult, WebsiteData, CompanyURL, PostedWebsite,TaskInfo
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.http import HttpResponse, JsonResponse
from celery.result import AsyncResult
import logging
from django.views.decorators.http import require_http_methods
from django.http import FileResponse, HttpResponseNotFound,Http404
from django.conf import settings
import json
import os
import glob
from urllib.parse import urlparse
from django.core.exceptions import ValidationError
from django.http import HttpResponseRedirect
from django.urls import reverse
from .task_content_gen import process_xls_file_incrementally
import tempfile
from django.views.decorators.cache import never_cache



logger = logging.getLogger(__name__)


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            # Redirect to a success page.
            return redirect('home')  # Replace 'home' with the name of your target page
        else:
            # Return an 'invalid login' error message.
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'registration/login.html')  # Replace with your template name

@login_required
def stop_process(request):
    request.session['stop_signal'] = True
    return JsonResponse({'status': 'stopped'})

def home(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        media_dir = os.path.join(settings.MEDIA_ROOT, 'generated_files')
        if not os.path.exists(media_dir):
            os.makedirs(media_dir)

        # Clear the session value before uploading a new file
        if 'uploaded_file_name' in request.session:
            del request.session['uploaded_file_name']
        # Clear the session value before uploading a new file
        if 'row_values' in request.session:
            del request.session['row_values']
            
        # Get the uploaded file
        excel_file = request.FILES['excel_file']
        uploaded_file_name = excel_file.name  # Get the uploaded file's name
        
        # Store the file name in session
        request.session['uploaded_file_name'] = uploaded_file_name

        # Open the Excel file
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        second_row = sheet[2]
        row_values = [cell.value for cell in second_row]

        # Store the row_values in the session
        request.session['row_values'] = row_values

        # Continue processing the rest of your logic
        company_website = row_values[4]

        api_configs_count = APIConfig.objects.filter(site_enable=True).count()
        site_number = int(request.POST.get('site_number'))

        if api_configs_count < site_number:
            return JsonResponse({
                'error': 'Not enough websites to run the requested number of tasks.',
                'api_configs_count': api_configs_count,
            }, status=400)

        api_configs = APIConfig.objects.filter(site_enable=True).order_by('?')
        GeneratedURL.objects.filter(user=request.user).delete()
        task_ids = []
        delay_seconds = 0  # initial delay
        processed_sites_count = 0
        api_configs_iterator = iter(api_configs)

        match_root_domain = 'match_root_domain' in request.POST

        while processed_sites_count < site_number:
            try:
                api_config = next(api_configs_iterator)

                # Handle WebsiteData for duplication check
                website_data, _ = WebsiteData.objects.get_or_create(api_config=api_config)
                
                # Check for existing company websites
                existing_company_websites = json.loads(website_data.company_websites) if website_data.company_websites else []

                if match_root_domain:
                    new_company_website_root = get_root_domain(company_website)
                    existing_roots = [get_root_domain(url) for url in existing_company_websites]
                    if new_company_website_root in existing_roots:
                        task_ids.append('skipped_task_' + str(processed_sites_count))
                        continue
                else:
                    if company_website in existing_company_websites:
                        task_ids.append('skipped_task_' + str(processed_sites_count))
                        continue
                            
                website = api_config.website.rstrip('/')
                json_url = f"https://{website}/wp-json/wp/v2"
                task = create_company_profile_post.apply_async(
                    args=[row_values, json_url, api_config.website, api_config.user, api_config.password, api_config.template_no],
                    countdown=delay_seconds
                )
                task_ids.append(task.id)
                delay_seconds += 2  # Increment the delay for the next task

            except StopIteration:
                # If no more API configurations to process, add a placeholder task ID
                task_ids.append('placeholder_task_' + str(processed_sites_count))
            
            processed_sites_count += 1  # Increment the count of processed sites

        return JsonResponse({'task_ids': task_ids})   

    return render(request, "listing/index.html")



@login_required
def get_task_result(request, task_id):
    task_result = AsyncResult(task_id)
    if task_result.ready():
        try:
            Author_name, url, website, company_website = task_result.result
            if url:
                logger.info(f"URL saved to database: {url}")
                GeneratedURL.objects.create(user=request.user, url=url, author_name=Author_name)
                CompanyURL.objects.create(generated_url=url, company_website=company_website)

                # Retrieve or create WebsiteData for the website
                api_config = APIConfig.objects.get(website=website)
                website_data, created = WebsiteData.objects.get_or_create(api_config=api_config)

                # Handle company_websites
                existing_company_websites = json.loads(website_data.company_websites) if website_data.company_websites else []
                if company_website not in existing_company_websites and company_website != "":
                    existing_company_websites.append(company_website)
                    website_data.company_websites = json.dumps(existing_company_websites)

                # Save the updated website_data
                website_data.save()

                return JsonResponse({'status': 'SUCCESS', 'url': url, 'author_name': Author_name})
            else:
                logger.warning("Task returned an empty URL.")
                failure_url = f"{website} - Failed to Post in this Domain" if website else "Task Failed"
                GeneratedURL.objects.create(user=request.user, url=failure_url)
                return JsonResponse({'status': 'FAILURE', 'error': 'Empty URL'})
        except Exception as e:
            logger.error(f"Error in get_task_result in {website}: {e}", exc_info=True)
            return JsonResponse({'status': 'ERROR', 'error': str(e)})
    else:
        return JsonResponse({'status': 'PENDING'})

@login_required
def unique_consecutive_domain(request):
    if request.method == 'POST' and request.FILES.get('excel_file', None):
        excel_file = request.FILES['excel_file']
        
        
         # Clear the session value before uploading a new file
        if 'uploaded_file_name' in request.session:
            del request.session['uploaded_file_name']
        # Clear the session value before uploading a new file
        if 'row_values' in request.session:
            del request.session['row_values']
            
        uploaded_file_name = excel_file.name
        request.session['uploaded_file_name'] = uploaded_file_name

        site_number = int(request.POST.get('site_number'))

        already_posted_websites = PostedWebsite.objects.filter(user=request.user).values_list('website__website', flat=True)
        print(f"Already posted websites: {list(already_posted_websites)}")

        available_api_configs = APIConfig.objects.filter(site_enable=True).exclude(website__in=already_posted_websites)
        print(f"Available API configs: {list(available_api_configs.values_list('website', flat=True))}")

        total_available_sites = available_api_configs.count()
        shortage = site_number - total_available_sites
        print(f"Total available sites: {total_available_sites}, Shortage: {shortage}")

        if shortage > 0:           
            additional_websites_qs = PostedWebsite.objects.filter(                
                website__site_enable=True  # Ensure this condition correctly refers to APIConfig's site_enable field
            ).order_by('?')[:shortage]

            additional_websites = additional_websites_qs.values_list('website__website', flat=True)
            print(f"Additional websites from PostedWebsite: {list(additional_websites)}")

            additional_api_configs = APIConfig.objects.filter(website__in=additional_websites)
            all_api_configs = available_api_configs.union(additional_api_configs)
            print("ALL API CONFIG SITE afte union:", all_api_configs)
        else:
            all_api_configs = available_api_configs

        print(f"All API configs to use: {list(all_api_configs.values_list('website', flat=True))}")


        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        second_row = sheet[2]
        row_values = [cell.value for cell in second_row]
        
        # Store the row_values in the session
        request.session['row_values'] = row_values
        company_website = row_values[4]

        GeneratedURL.objects.filter(user=request.user).delete()
        task_ids = []
        delay_seconds = 0

        for api_config in all_api_configs[:site_number]:
            # Ensure unique company websites if match_root_domain is required
            match_root_domain = 'match_root_domain' in request.POST
            if match_root_domain:
                new_company_website_root = get_root_domain(company_website)
                website_data, _ = WebsiteData.objects.get_or_create(api_config=api_config)
                existing_roots = [get_root_domain(url) for url in json.loads(website_data.company_websites or "[]")]
                if new_company_website_root in existing_roots:
                    continue

            print("Creating task for website:", api_config.website)
            task = create_company_profile_post.apply_async(
                args=[row_values, f"https://{api_config.website}/wp-json/wp/v2", api_config.website, api_config.user, api_config.password, api_config.template_no],
                countdown=delay_seconds
            )
            task_ids.append(task.id)
            delay_seconds += 4

        if not task_ids:
            messages.error(request, 'No tasks were scheduled due to configuration or eligibility issues.')
            return redirect('unique_consecutive_domain')  
        # Before returning JsonResponse
        if shortage > 0:
            used_websites_names = additional_api_configs.values_list('website', flat=True)
            additional_message = f"Note: Due to a shortage of available new websites, some tasks are scheduled on previously used websites: {', '.join(used_websites_names)}."
        else:
            additional_message = ''

        return JsonResponse({
            'message': 'Tasks scheduled successfully.',
            'task_ids': task_ids,
            'used_posted_websites': shortage > 0,
            'additional_message': additional_message,
        })
    else:
        return render(request, "listing/unique_domain.html")  


@login_required
def get_task_result_unique(request, task_id):
    task_result = AsyncResult(task_id)
    if task_result.ready():
        try:
            Author_name, url, website, company_website = task_result.result
            if url:
                logger.info(f"URL saved to database: {url}")
                # Save the generated URL and associated data
                GeneratedURL.objects.create(user=request.user, url=url, author_name=Author_name)
                CompanyURL.objects.create(generated_url=url, company_website=company_website)

                # Retrieve or create the APIConfig for the website
                try:
                    api_config = APIConfig.objects.get(website=website)
                    
                    # Save to PostedWebsite if not already saved
                    posted_website, created = PostedWebsite.objects.get_or_create(website=api_config, user=request.user)
                    if created:
                        logger.info(f"New PostedWebsite entry created for {website}")
                    else:
                        logger.info(f"PostedWebsite entry already exists for {website}")

                    # Update WebsiteData with company website if not already included
                    website_data, _ = WebsiteData.objects.get_or_create(api_config=api_config)
                    existing_company_websites = json.loads(website_data.company_websites) if website_data.company_websites else []
                    if company_website not in existing_company_websites and company_website != "":
                        existing_company_websites.append(company_website)
                        website_data.company_websites = json.dumps(existing_company_websites)
                        website_data.save()
                
                except APIConfig.DoesNotExist:
                    logger.error(f"APIConfig does not exist for website: {website}")

                return JsonResponse({'status': 'SUCCESS', 'url': url, 'author_name': Author_name})
            else:
                logger.warning("Task returned an empty URL.")
                failure_url = f"{website} - Failed to Post in this Domain" if website else "Task Failed"
                GeneratedURL.objects.create(user=request.user, url=failure_url)
                return JsonResponse({'status': 'FAILURE', 'error': 'Empty URL'})
        except Exception as e:
            logger.error(f"Error in get_task_result for {website}: {e}", exc_info=True)
            return JsonResponse({'status': 'ERROR', 'error': str(e)})
    else:
        return JsonResponse({'status': 'PENDING'})
    



@login_required
def get_generated_links_json(request):
    # Ensure that the row_values are present in the session
    row_values = request.session.get('row_values', None)

    if not row_values:
        return JsonResponse({'error': 'No row values found in the session.'}, status=400)

    # Extract company name and description from the session-stored row_values
    company_name = row_values[0]  # Assuming company name is at index 0
    description = row_values[2]  # Assuming description is at index 2

    # Fetch the generated URLs for the current user
    links_data = GeneratedURL.objects.filter(user=request.user).order_by('created_at').values('url', 'author_name')
    links_list = [{'url': link['url'], 'author_name': link['author_name']} for link in links_data]

    # Before writing the Excel file, post the summary to WordPress and get the published URL
    live_urls = [link['url'] for link in links_data]  # Use the list of URLs from the generated links
    post_url = post_summary_to_wordpress(company_name, description, live_urls)

    # Check if the post was successfully published
    if post_url:
        logger.info(f"Summary post published successfully: {post_url}")
    else:
        logger.error("Failed to create and publish the summary post.")
        post_url = "Failed to publish"  # If the post fails, indicate it in the Excel file

    # Create a new workbook to save the generated links
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    uploaded_file_name = request.session.get('uploaded_file_name', 'Generated Links')
    base_file_name, _ = os.path.splitext(uploaded_file_name)
    base_file_name += "-links"
    sanitized_base_name = ''.join(char for char in base_file_name if char.isalnum() or char in " -_")
    sanitized_base_name = sanitized_base_name[:31]  # Excel sheet title cannot exceed 31 characters
    sheet.title = sanitized_base_name
    
    # Ensure the file name is unique
    file_name = f"{sanitized_base_name}.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, 'generated_files', file_name)

    counter = 1
    while os.path.exists(file_path):
        file_name = f"{sanitized_base_name}-{counter}.xlsx"
        file_path = os.path.join(settings.MEDIA_ROOT, 'generated_files', file_name)
        counter +=1

    # Set column headers
    sheet['A1'] = 'SL Number'
    sheet['B1'] = 'Root Domain'
    sheet['C1'] = 'URL'
    sheet['D1'] = 'Author Name'

    # Write the generated links to the Excel file
    for index, link in enumerate(links_data, start=2):
        url = link['url']
        author_name = link['author_name']
        root_domain = urlparse(url).netloc
        sheet.cell(row=index, column=1, value=index-1)
        sheet.cell(row=index, column=2, value=root_domain)
        sheet.cell(row=index, column=3, value=url)
        sheet.cell(row=index, column=4, value=author_name)

    # After writing the links, add the "Summary URL"
    sheet.cell(row=index + 1, column=2, value="Summary URL:")
    sheet.cell(row=index + 1, column=3, value=post_url)  # Add the post URL

    try:
        workbook.save(file_path)
        logger.info(f"Excel file successfully saved at {file_path}")
        # Save the unique file name in the session for later retrieval
        request.session['download_file_name'] = file_name
        
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create Excel file'})

    # Provide the URL to download the Excel file
    file_url = request.build_absolute_uri(os.path.join(settings.MEDIA_URL, 'generated_files', file_name))
    
    # Return the Excel file URL and the generated links as well as the summary URL
    return JsonResponse({
        'excel_file_url': file_url,
        'links': links_list,
        'post_url': post_url  # Add the post_url to the response
    })





@login_required
def download_excel_latest(request):
    download_file_name = request.session.get('download_file_name')
    if not download_file_name:
        return HttpResponseNotFound('No file name found in the session.')

    file_path = os.path.join(settings.MEDIA_ROOT, 'generated_files', download_file_name)

    if os.path.exists(file_path):
        # Open the file without using 'with' to prevent auto-closing
        file = open(file_path, 'rb')
        response = FileResponse(file, as_attachment=True, filename=download_file_name)
        # No need to manually close the file; FileResponse will handle it.
        return response
    else:
        return HttpResponseNotFound('The requested file was not found on our server.')



@login_required
def get_django_messages(request):
    messages_list = []
    for message in messages.get_messages(request):
        messages_list.append({
            "level": message.level,
            "message": message.message,
            "tags": message.tags,
        })
    return JsonResponse(messages_list, safe=False)


def site_data(request):
    if request.method == 'POST':
        excel_file = request.FILES['site_excel_file']

        # Read the Excel file
        df = pd.read_excel(excel_file)

        # Iterate over the rows and update the database
        for index, row in df.iterrows():
            APIConfig.objects.update_or_create(
                website=row['url'],
                defaults={
                    'user': row['user'],
                    'password': row['password'],
                    'template_no': row['template_no']
                }
            )
        
        # Add a success message
        messages.success(request, "File uploaded and database updated.")
        return redirect('site_data')  # Redirect to the same page

    return render(request, "listing/site_data.html")


def get_api_config_data(request):
    data = list(APIConfig.objects.values('website', 'user', 'password', 'template_no'))
    return JsonResponse(data, safe=False)

@require_http_methods(["GET", "POST"])
def rest_api_test(request):
    context = {'api_configs': APIConfig.objects.all()}

    # Clear previous results when the page is freshly loaded (GET request)
    if request.method == 'GET':
        TestResult.objects.all().delete()
        request.session['test_status'] = {}  # Also clear the session info if you are using it

    if request.method == 'POST':
        # If the 'Test All' button is pressed
        if 'test_all' in request.POST:
            task_ids = []
            for config in context['api_configs']:
                perform_test_task.delay(config.id)

            # Store the task_ids in the session for checking completion
            request.session['test_all_task_ids'] = [perform_test_task.delay(config.id).id for config in context['api_configs']]
            
            # Redirect to avoid re-posting on refresh
            return redirect('rest_api_test')

        # If the 'Test Site' button is pressed for a single site
        elif 'test_single' in request.POST:
            selected_url = request.POST.get('api_url')
            try:
                selected_config = APIConfig.objects.get(website=selected_url)
                perform_test_task.delay(selected_config.id)
            except APIConfig.DoesNotExist:
                messages.error(request, "Selected site configuration does not exist.")
            except APIConfig.MultipleObjectsReturned:
                messages.error(request, "Multiple configurations found for the selected site.")
            except Exception as e:
                messages.error(request, f"An error occurred on {selected_url}: {e}")

        # Redirect to avoid re-posting on refresh
        return redirect('rest_api_test')

    # Render the page with context
    return render(request, 'listing/rest_api_test.html', context)


def test_status_update(request):
    # Fetch test results and prepare the context
    test_results = TestResult.objects.all()
    test_status = {result.config.website: result.status for result in test_results}

    failed_urls = [(url, status.split(":")[1].strip()) for url, status in test_status.items() if 'Failed' in status]

    # Always create an Excel file, but it will be empty if there are no failed URLs
    try:
        excel_file = os.path.join(settings.BASE_DIR, 'failed_tests.xlsx')
        logger.info(f"Creating Excel file at: {excel_file}")

        # Create a DataFrame, it will be empty if failed_urls is empty
        new_data = pd.DataFrame(failed_urls, columns=['URL', 'Status'])
        new_data.to_excel(excel_file, index=False)

        if not failed_urls:
            logger.info("No failed URLs to save, creating an empty Excel file.")

    except Exception as e:
        logger.error(f"Error while creating Excel file: {e}")

    return JsonResponse(test_status)


def perform_test(request, config):
    try:
        username = config.user
        password = config.password
        response = test_post_to_wordpress(config.website, username, password, "Test Content")
        test_status = request.session.get('test_status', {})
        if response.status_code in [201]:
            test_status[config.website] = 'Success: Post Created successfully'                
            json_data = response.json()
            post_link = json_data.get('link', None)
            messages.success(request, f"Post Created successfully on {post_link}.")
            
            response_data = json.loads(response.text)
            post_id = response_data.get('id')
            delete_response = delete_from_wordpress(config.website, username, password, post_id)
            
            if delete_response is not None and delete_response.status_code == 200:
                messages.success(request, f"Post deleted successfully on {post_link}.")
            else:
                messages.error(request, "Failed to delete post.")
        else:
            test_status[config.website] = f'Failed: Site test failed with status code: {response.status_code}'
            request.session['test_status'] = test_status
            messages.error(request, f"Site test failed with status code: {response.status_code}")
            # Save the failed URL to an Excel file            
    except Exception as e:
        messages.error(request, f"An error occurred while testing {config.website}: {e}")


def download_failed_list(request):
    file_name = 'failed_tests.xlsx'
    file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), file_name)

    # Check if the file exists
    if not os.path.exists(file_path):
        messages.error(request, "Failed Tests file not found.")
        return HttpResponseNotFound('<h1>File not found</h1>')

    # Open the file without using a context manager
    f = open(file_path, 'rb')
    response = FileResponse(f)
    response['Content-Disposition'] = 'attachment; filename="failed_tests.xlsx"'
    return response


        
def list_files(request):
    media_subdir = os.path.join(settings.MEDIA_ROOT, 'generated_files')
    # List all .xlsx files along with their full path
    files = [os.path.join(media_subdir, f) for f in os.listdir(media_subdir) if f.endswith('.xlsx')]
    
    # Sort files by modification time in descending order
    files.sort(key=os.path.getmtime, reverse=True)
    
    # Extract file names from the sorted list of file paths
    file_names = [os.path.basename(f) for f in files]
    
    return JsonResponse(file_names, safe=False)

def download_file(request):
    file_name = request.GET.get('file')  # Get the file name from request
    file_path = os.path.join(settings.MEDIA_ROOT, 'generated_files', file_name)

    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
    raise Http404  # Return 404 if file not found

def delete_all_files(request):
    media_subdir = os.path.join(settings.MEDIA_ROOT, 'generated_files')
    files = glob.glob(os.path.join(media_subdir, '*.xlsx'))
    deleted_files, failed_files = [], []

    for f in files:
        try:
            os.remove(f)
            deleted_files.append(f)
        except Exception as e:
            failed_files.append((f, str(e)))

    if failed_files:
        # Return a response indicating which deletions failed
        return JsonResponse({
            'status': 'partial_success',
            'deleted_files': deleted_files,
            'failed_files': failed_files
        })
    return JsonResponse({'status': 'success', 'deleted_files': deleted_files})


@login_required
def delete_posts(request):
    deleted_posts_urls = request.session.pop('deleted_posts_urls', []) if 'deleted_posts_urls' in request.session else []

    if request.method == 'POST':
        links = []  # Initialize links list
        excel_file = request.FILES.get('excel_file')
        links_textarea = request.POST.get('links')
        TaskInfo.objects.all().delete()

        if excel_file:
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                if row[0]:  # Check if the cell is not None
                    link = row[0].strip()  # Strip spaces
                    if not link.endswith('/'):  # Check if the link ends with a slash
                        link += '/'  # Add a slash if it doesn't
                    links.append(link)
        elif links_textarea:
            links = [link.strip() for link in links_textarea.splitlines()]  # List comprehension to strip spaces
            links = [link + '/' if not link.endswith('/') else link for link in links]  # Ensure each link ends with a slash

        if links:
            messages.success(request, 'Deletion process started for submitted posts. Check deletion logs for status.')
        else:
            messages.info(request, 'No valid links provided for deletion.')
        for post_url in links:
            task = delete_post_by_url.delay(post_url)  # Initiates the task
            TaskInfo.objects.create(post_url=post_url, task_id=task.id)  # Save task info



        # Update the session with URLs attempted for deletion
        request.session['deleted_posts_urls'] = deleted_posts_urls

        # Redirect to avoid POST data resubmission
        return redirect('delete_posts')

    # For GET request or after the POST processing, render the page with the context
    context = {'deleted_posts_urls': deleted_posts_urls}
    return render(request, 'listing/delete_posts.html', context)


def check_delete_task_status(request):
    task_info_objs = TaskInfo.objects.all()
    results = []
    all_tasks_completed = True  # Assume all tasks are completed initially

    for info in task_info_objs:
        task_id = info.task_id
        task_result = AsyncResult(task_id)

        if task_result.ready():
            success = task_result.result  # This will be True or False as returned by the task
            info.status = 'SUCCESS' if success else 'FAILED'
            info.save()  # Update the status in the database
        else:
            all_tasks_completed = False  # If any task isn't ready, mark as not all completed

        results.append({
            'url': info.post_url,
            'status': info.status,
        })

    return JsonResponse({
        'tasks': results,
        'all_tasks_completed': all_tasks_completed  # Add this to signal if all tasks are done
    })
    
@login_required
def flash_posted_website(request):
    if request.method == "POST":
        if PostedWebsite.objects.count() == 0:
            # If there are no objects, flash a message indicating the database is already empty
            messages.info(request, "The database is already empty.")
        else:
            # Delete all objects in PostedWebsite if there are any
            PostedWebsite.objects.all().delete()
            messages.success(request, "All data in Posted Website Records have been successfully deleted.")
    else:
        messages.error(request, "Invalid request method.")
    
    # Redirect to a certain page after the action
    return HttpResponseRedirect(reverse('unique_consecutive_domain'))


@login_required
def post_update_view(request):
    logger.info("Entered post_update_view")
    messages_list = []

    if request.method == 'POST':
        post_urls = request.POST.get('post_urls').splitlines()
        excel_file = request.FILES.get('excel_file')

        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
                second_row = sheet[2]
                row_values = [cell.value for cell in second_row if cell.value is not None]
            except Exception as e:
                logger.error(f"Failed to read Excel file: {str(e)}")
                messages_list.append({'tags': 'error', 'message': f"Failed to read Excel file: {str(e)}"})
                return render(request, 'listing/post_update.html', {'messages': messages_list})

            task_ids = []

            for post_url in post_urls:
                parsed_url = urlparse(post_url)
                domain = parsed_url.netloc

                try:
                    config = APIConfig.objects.get(website__icontains=domain)
                except APIConfig.DoesNotExist:
                    logger.error(f"No API configuration found for domain: {domain}")
                    messages_list.append({'tags': 'error', 'message': f"No API configuration found for domain: {domain}"})
                    continue

                post_id = find_post_id_by_url(domain, post_url, config.user, config.password)
                if post_id is None:
                    logger.error(f"Post with URL '{post_url}' not found.")
                    messages_list.append({'tags': 'error', 'message': f"Post with URL '{post_url}' not found."})
                    continue

                json_url = f"https://{domain}/wp-json/wp/v2/posts/{post_id}"

                # Trigger the task asynchronously using delay() or apply_async()
                task = update_company_profile_post.delay(
                    row_values=row_values,
                    json_url=json_url,
                    website=config.website,
                    user=config.user,
                    password=config.password,
                    html_template=config.template_no
                )
                task_ids.append(task.id)

            messages_list.append({'tags': 'info', 'message': f"Tasks started. Task IDs: {task_ids}"})

            logger.info(f"All tasks queued. Task IDs: {task_ids}")
            return render(request, 'listing/post_update.html', {
                'messages': messages_list,
                'task_ids': task_ids  
            })

    return render(request, 'listing/post_update.html')



@login_required
def check_task_status_post_update(request, task_id):
    logger.info(f"Checking status for task: {task_id}")
    task_result = AsyncResult(task_id)

    if task_result.state == 'PENDING':
        logger.info(f"Task {task_id} is still pending")
        return JsonResponse({
            'task_id': task_id,
            'state': 'PENDING',
            'info': 'Task is queued or waiting to be picked up by a worker.'
        })
    elif task_result.state == 'STARTED':
        logger.info(f"Task {task_id} has started")
        return JsonResponse({
            'task_id': task_id,
            'state': 'STARTED',
            'info': 'Task has been picked up by a worker and is currently running.'
        })
    elif task_result.state == 'RETRY':
        logger.info(f"Task {task_id} is being retried")
        return JsonResponse({
            'task_id': task_id,
            'state': 'RETRY',
            'info': 'Task encountered an error and is being retried.'
        })
    elif task_result.state == 'FAILURE':
        logger.error(f"Task {task_id} failed: {task_result.result}")
        return JsonResponse({
            'task_id': task_id,
            'state': 'FAILURE',
            'info': str(task_result.result)
        })
    elif task_result.state == 'SUCCESS':
        logger.info(f"Task {task_id} completed successfully: {task_result.result}")
        result = task_result.result
        return JsonResponse({
            'task_id': task_id,
            'state': 'SUCCESS',
            'info': result.get('message', 'Task completed successfully.'),
            'status': result.get('status', 'success')
        })
    else:
        logger.warning(f"Task {task_id} is in an unexpected state: {task_result.state}")
        return JsonResponse({
            'task_id': task_id,
            'state': task_result.state,
            'info': 'Task is in an unexpected state.'
        })
@login_required
def content_gen_view(request):
    if request.method == 'POST':
        xlsx_file = request.FILES.get('xlsx_file')
        num_sites = request.POST.get('num_sites', '1')
        avoid_root_domain = request.POST.get('avoid_root_domain') == 'on'  # Check if the checkbox is checked

        try:
            num_sites = int(num_sites)
            if num_sites < 1:
                return JsonResponse({'status': 'error', 'message': 'Number of sites must be at least 1.'}, status=400)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid number of sites.'}, status=400)

        if xlsx_file:
            try:
                with tempfile.NamedTemporaryFile(suffix='.xlsx', dir='/app/tempfiles', delete=False) as temp_file:
                    for chunk in xlsx_file.chunks():
                        temp_file.write(chunk)
                    temp_file_path = temp_file.name

                # Log the file path for debugging
                logger.info(f"Temporary file created at: {temp_file_path}")

                # You need to pass original_filename to the task
                original_filename = xlsx_file.name.rsplit('.', 1)[0]
                # Call the Celery task
                task = process_xls_file_incrementally.delay(temp_file_path, num_sites, avoid_root_domain, original_filename)

                # Return a response indicating the task has started
                return JsonResponse({
                    'status': 'in_progress',
                    'task_id': task.id,
                    'message': 'File is being processed. Check the status later.'
                })
            except Exception as e:
                logger.error(f"An error occurred: {str(e)}")
                return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'}, status=500)
        else:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    # If it's a GET request, render the content generation page
    return render(request, 'listing/content_generation.html')



@never_cache
@login_required
def content_gen_status_view(request, task_id):
    result = AsyncResult(task_id)
    logger.info(f"Checking status for task ID: {task_id} - Current state: {result.state}")

    try:
        if result.state == 'PENDING':
            response = {
                'status': 'in_progress',
                'message': 'Task is pending...',
            }
        elif result.state == 'SUCCESS':
            # Access the result from the meta dictionary
            meta_data = result.info  # This should contain your meta information
            urls = meta_data.get('result', []) if isinstance(meta_data.get('result'), list) else []
            successful_postings = meta_data.get('successful_postings', 0)
            file_path = meta_data.get('file_path', '')

            response = {
                'status': 'SUCCESS',
                'message': 'Task completed successfully.',
                'result': urls,  # This is now correctly referring to the URLs
                'successful_postings': successful_postings,
                'file_path': file_path
            }

        elif result.state == 'FAILURE':
            exc_info = result.info
            error_message = str(exc_info) if exc_info else 'Task failed with no error message available'
            response = {
                'status': 'FAILURE',
                'message': error_message,
                'exception_type': type(exc_info).__name__ if exc_info else 'Unknown'
            }
        elif result.state == 'single_post_done':
            # Check if result.info has the necessary metadata
            meta_info = result.info if isinstance(result.info, dict) else {}
            response = {
                'status': 'single_post_done',
                'meta': meta_info,  # This should contain current, total, and last_posted_url
                'message': 'Single post completed',
            }
        else:
            response = {
                'status': result.state,
                'message': 'Unknown task state.',
            }
    except Exception as e:
        logger.error(f"Error retrieving task status: {type(e).__name__} - {str(e)}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'Failed to retrieve task status.'}, status=500)

    return JsonResponse(response)





@login_required
def list_files_geo(request):
    output_dir = os.path.join(settings.MEDIA_ROOT, 'GEO_output_folder')
    
    try:
        files = os.listdir(output_dir)
        
        if not files:
            return JsonResponse({'status': 'error', 'message': 'No files available.'}, status=404)

        return JsonResponse({'status': 'success', 'files': files})

    except FileNotFoundError:
        return JsonResponse({'status': 'error', 'message': 'Directory not found.'}, status=404)


@login_required
def download_file_geo(request, file_name):
    # Construct the full file path
    file_path = os.path.join(settings.MEDIA_ROOT, 'GEO_output_folder', file_name)
    
    # Check if the file exists
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename={file_name}'
            return response
    else:
        raise Http404("File not found.")

@login_required
def delete_all_files_geo(request):
    media_subdir = os.path.join(settings.MEDIA_ROOT, 'GEO_output_folder')
    files = glob.glob(os.path.join(media_subdir, '*.xls'))
    deleted_files, failed_files = [], []

    for f in files:
        try:
            os.remove(f)
            deleted_files.append(os.path.basename(f))  # Append just the file name
        except Exception as e:
            failed_files.append((os.path.basename(f), str(e)))

    if failed_files:
        return JsonResponse({
            'status': 'partial_success',
            'deleted_files': deleted_files,
            'failed_files': failed_files
        })
    
    return JsonResponse({'status': 'success', 'deleted_files': deleted_files})
